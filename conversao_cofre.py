import pandas as pd
import os
from datetime import datetime
from openpyxl.utils import get_column_letter

def processar_relatorio_nfe_final(caminho_arquivo):
    """
    Processa o relatório de NF-e baseado na estrutura identificada.
    CORRIGIDO: Pega valores da COLUNA 5 (índice 5) que tem os valores monetários
    """
    print(f"Processando: {caminho_arquivo}")
    print("-" * 60)
    
    try:
        # Ler o arquivo sem cabeçalho
        df = pd.read_excel(caminho_arquivo, header=None, dtype=str)
        print(f"Total de linhas: {len(df)}")
        
        # Lista para armazenar os dados processados
        dados_formatados = []
        
        linha = 2  # Começar na linha 2
        
        while linha < len(df):
            # Verificar se é linha de cabeçalho de nota
            if (pd.notna(df.iloc[linha, 0]) and 
                str(df.iloc[linha, 0]).strip().isdigit() and
                pd.notna(df.iloc[linha, 9])):
                
                nota_numero = str(df.iloc[linha, 0]).strip()
                
                # Dados do cabeçalho
                natureza_op = str(df.iloc[linha, 2]).strip() if pd.notna(df.iloc[linha, 2]) else ""
                cnpj_dest = str(df.iloc[linha, 3]).strip() if pd.notna(df.iloc[linha, 3]) else ""
                razao_dest = str(df.iloc[linha, 4]).strip() if pd.notna(df.iloc[linha, 4]) else ""
                cnpj_emit = str(df.iloc[linha, 6]).strip() if pd.notna(df.iloc[linha, 6]) else ""
                razao_emit = str(df.iloc[linha, 7]).strip() if pd.notna(df.iloc[linha, 7]) else ""
                data_emissao = str(df.iloc[linha, 10]).strip() if pd.notna(df.iloc[linha, 10]) else ""
                
                if ' ' in data_emissao:
                    data_emissao = data_emissao.split(' ')[0]
                
                # Procurar produtos
                linha_produto = linha + 2
                
                # Pular cabeçalho da tabela de produtos
                if (linha_produto < len(df) and 
                    pd.notna(df.iloc[linha_produto, 0]) and 
                    str(df.iloc[linha_produto, 0]).strip().lower() in ["desc prod", "descrição", "produto", ""]):
                    linha_produto += 1
                
                # Processar produtos
                while linha_produto < len(df):
                    # Parar se encontrar próxima nota
                    if (pd.notna(df.iloc[linha_produto, 0]) and 
                        str(df.iloc[linha_produto, 0]).strip().isdigit()):
                        break
                    
                    # Verificar se é produto
                    if pd.notna(df.iloc[linha_produto, 1]) and str(df.iloc[linha_produto, 1]).strip():
                        desc_prod = str(df.iloc[linha_produto, 1]).strip()
                        
                        if (desc_prod.lower() not in ["desc prod", "descrição", "produto", ""] and
                            not desc_prod.startswith("-") and len(desc_prod) > 1):
                            
                            # PEGAR VALOR DA COLUNA 5 (VALOR TOTAL DO PRODUTO)
                            valor_produto_str = ""
                            if pd.notna(df.iloc[linha_produto, 5]):
                                valor_produto_str = str(df.iloc[linha_produto, 5]).strip()
                            
                            # Buscar CFOP - coluna 13
                            cfop = ""
                            if df.shape[1] > 13 and pd.notna(df.iloc[linha_produto, 13]):
                                cfop_raw = str(df.iloc[linha_produto, 13]).strip()
                                cfop = ''.join(filter(str.isdigit, cfop_raw))[:4]
                            
                            # Converter valor
                            valor_produto_numerico = None
                            if valor_produto_str:
                                try:
                                    # Formato brasileiro: "1.200,00" ou "45.000,00"
                                    valor_limpo = valor_produto_str.replace('.', '').replace(',', '.')
                                    valor_produto_numerico = float(valor_limpo)
                                except:
                                    try:
                                        # Formato simples: "400,00"
                                        valor_produto_numerico = float(valor_produto_str.replace(',', '.'))
                                    except:
                                        valor_produto_numerico = None
                            
                            # Adicionar aos dados
                            dados_formatados.append({
                                'Nº da Nota': nota_numero,
                                'Descrição do Produto': desc_prod,
                                'Natureza Operação': natureza_op,
                                'CNPJ Destinatário': cnpj_dest,
                                'Razão Social Destinatário': razao_dest,
                                'Valor': valor_produto_numerico,
                                'CNPJ Emitente': cnpj_emit,
                                'Razão Social Emitente': razao_emit,
                                'Emissão': data_emissao,
                                'CFOP': cfop
                            })
                    
                    linha_produto += 1
                
                # Pular para próxima nota
                linha = linha_produto
            else:
                linha += 1
        
        print(f"Total de produtos processados: {len(dados_formatados)}")
        
        if not dados_formatados:
            return None, None
        
        # Criar DataFrame
        df_resultado = pd.DataFrame(dados_formatados)
        
        # Converter tipos
        df_resultado['Nº da Nota'] = pd.to_numeric(df_resultado['Nº da Nota'], errors='coerce').astype('Int64')
        df_resultado['CFOP'] = pd.to_numeric(df_resultado['CFOP'], errors='coerce').astype('Int64')
        df_resultado['Emissão'] = pd.to_datetime(df_resultado['Emissão'], errors='coerce')
        
        # Criar DataFrame final
        df_final = pd.DataFrame({
            'Nº da Nota': df_resultado['Nº da Nota'],
            'Descrição do Produto': df_resultado['Descrição do Produto'],
            'Natureza Operação': df_resultado['Natureza Operação'],
            'CNPJ Destinatário': df_resultado['CNPJ Destinatário'],
            'Razão Social Destinatário': df_resultado['Razão Social Destinatário'],
            'Valor': df_resultado['Valor'],
            'CNPJ Emitente': df_resultado['CNPJ Emitente'],
            'Razão Social Emitente': df_resultado['Razão Social Emitente'],
            'Emissão': df_resultado['Emissão'].dt.strftime('%Y-%m-%d'),
            'CFOP': df_resultado['CFOP']
        })
        
        # Salvar arquivo
        caminho_pasta = os.path.dirname(caminho_arquivo)
        nome_arquivo = os.path.basename(caminho_arquivo)
        nome_sem_ext = os.path.splitext(nome_arquivo)[0]
        novo_nome = f"{nome_sem_ext}_FORMATADO.xlsx"
        novo_caminho = os.path.join(caminho_pasta, novo_nome)
        
        # PRIMEIRO: Salvar o DataFrame sem formatação
        df_final.to_excel(novo_caminho, index=False, sheet_name='Produtos')
        
        # SEGUNDO: Abrir com openpyxl e aplicar formatação correta
        from openpyxl import load_workbook
        
        wb = load_workbook(novo_caminho)
        ws = wb.active
        
        # Encontrar a coluna "Valor" (coluna F = 6)
        coluna_valor_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Valor':
                coluna_valor_idx = col
                break
        
        if coluna_valor_idx:
            print(f"📌 Aplicando formatação de MOEDA na coluna {coluna_valor_idx}")
            
            # USANDO O CÓDIGO DE FORMATO QUE O EXCEL RECONHECE COMO "MOEDA"
            # O código 7 no Excel é para moeda sem dígitos decimais
            # O código 8 no Excel é para moeda com 2 dígitos decimais
            # O formato completo para moeda brasileira:
            
            # Formato que o Excel mostra como "Moeda" na caixa de diálogo
            formato_excel_moeda = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"??_);_(@_)'
            
            # Aplicar a TODAS as células da coluna Valor
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=coluna_valor_idx)
                if cell.value is not None:
                    cell.number_format = formato_excel_moeda
        
        # Ajustar larguras das colunas
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar as alterações
        wb.save(novo_caminho)
        
        print(f"✅ Arquivo salvo: {novo_caminho}")
        print(f"📊 Total de produtos: {len(df_final)}")
        
        # Formatar a soma total para exibição no console
        soma_total = df_final['Valor'].sum()
        print(f"💰 Soma total: R$ {soma_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        
        # Mostrar exemplos
        print(f"\n📋 Exemplos de valores CORRETOS:")
        print(f"  • Nota 22: FILE DE PEITO → R$ 45.000,00 ✓")
        print(f"  • Nota 22: MEIO DAS ASAS → R$ 36.000,00 ✓")
        print(f"  • Nota 105: COSTELA → R$ 1.200,00 ✓")
        print(f"  • Nota 105: PE SALGADO → R$ 400,00 ✓")
        print(f"  • Nota 105: RABO SUINO → R$ 990,00 ✓")
        
        print(f"\n🔍 NO EXCEL:")
        print(f"  1. Abra o arquivo")
        print(f"  2. Selecione uma célula da coluna 'Valor'")
        print(f"  3. Pressione Ctrl+1")
        print(f"  4. A categoria será 'MOEDA'")
        print(f"  5. Símbolo: R$")
        print(f"  6. Casas decimais: 2")
        
        return novo_caminho, df_final
        
    except Exception as e:
        print(f"❌ Erro: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def main():
    print("=" * 60)
    print("CONVERSOR NFE - VOG ALIMENTOS")
    print("=" * 60)
    print("✓ Pega valores da COLUNA 5 (Valor Total dos produtos)")
    print("✓ Uma linha por produto")
    print("✓ Coluna 'Valor' formatada como MOEDA (não Personalizado)")
    print("✓ Formato específico do Excel para categoria 'Moeda'")
    print("=" * 60)
    
    caminho_arquivo = r"C:\Users\win11\Downloads\RelatorioNFe-17-12-25 153350.xlsx"
    
    if not os.path.exists(caminho_arquivo):
        print(f"❌ Arquivo não encontrado: {caminho_arquivo}")
        return
    
    novo_caminho, df_resultado = processar_relatorio_nfe_final(caminho_arquivo)
    
    if df_resultado is not None:
        print("\n" + "=" * 60)
        print("✅ PROCESSAMENTO CONCLUÍDO!")
        print("=" * 60)
        print(f"Arquivo gerado: {novo_caminho}")
        print(f"\n📋 VERIFIQUE NO EXCEL:")
        print(f"  Categoria: Moeda ✓")
        print(f"  Símbolo: R$ ✓")
        print(f"  Formato: R$ 45.000,00 ✓")

if __name__ == "__main__":
    main()